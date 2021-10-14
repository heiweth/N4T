import mysql.connector
from mysql.connector import Error
from mysql.connector import errorcode
import csv
from datetime import date, timedelta, datetime
import time
import pandas as pd
from sqlalchemy import create_engine
import pymysql
import openpyxl
from styles_sanja import *
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from openpyxl import *
import os

path_to_scripts = "/home/cene-admin/n4t/"

today = date.today()
year = today.strftime("%Y")
year_today = today.strftime("%Y")



areas_not = ["PMUM", "SMP", "HUPX", "OPCOM", "OTE", "ELIX", "PHELIX"]
direction_to_insert = ["RSHU", "HURS", "MKRS", "RSMK", "RORS", "RSRO"]

datumi = ['2020-06-26','2020-06-27','2020-06-28','2020-06-29','2020-06-30','2020-07-01',	'2020-07-02',
	'2020-07-03',	'2020-07-04','2020-07-05','2020-07-06','2020-07-07','2020-07-08','2020-07-09',
	'2020-07-10','2020-07-11','2020-07-12','2020-07-13']

# area, date, hour, price
#sqlEngine = create_engine('mysql+pymysql://root:network@localhost/N4T')
sqlEngine = create_engine('mysql+pymysql://cene:Babastamena22@192.168.100.251/cene')
db_connection = sqlEngine.connect()

#tamnosiva
#ATHU HUAT HUHR HRHU HRRS RSHR SLOAT SIHR ATSLO HRSI GRIT ITGR CZDE(TenneT) DE(TenneT)CZ ATCZ CZAT

#zelena
#UA-BEI UA-IPS BRINDISI PUN SEEPEX CROPEX IBEX SMP PMUM EXAA SP EEX OTE OKTE OPCOM HUPX SKUA UASK HUUA UAHU ROUA UARO

#narandzasta
#HURS RSHU RORS RSRO BGRS RSBG MKRS RSMK

#base color
#ROBG BGRO

#peak color 
#BAHR HRBA MEBA BAME ALME MEAL ALGR GRAL GRTR TRGR MKGR GRMK

colour = {"BAHR":"fbe5d6" , "HRBA":"fbe5d6" , "MEBA":"fbe5d6" , "BAME":"fbe5d6" , "ALME":"fbe5d6" , "MEAL":"fbe5d6" , "ALGR":"fbe5d6" , "GRAL":"fbe5d6" , "GRTR":"fbe5d6" , "TRGR":"fbe5d6" , "MKGR":"fbe5d6" , "GRMK":"fbe5d6", "ATHU":"999999", "HUAT":"999999", "HUHR":"999999", "HRHU":"999999", "HRRS":"999999", "RSHR":"999999", "SLOAT":"999999", "SIHR":"999999", "ATSLO":"999999", "HRSI":"999999", "GRIT":"999999", "ITGR":"999999", "CZDE(TenneT)":"999999", "DE(TenneT)CZ":"999999", "ATCZ":"999999", "CZAT":"999999", "ROBG":"fff2cc", "BGRO":"fff2cc", "HURS":"fdb94d", "RSHU":"fdb94d", "RORS":"fdb94d", "RSRO":"fdb94d", "BGRS":"fdb94d", "RSBG":"fdb94d", "MKRS":"fdb94d", "RSMK":"fdb94d", "UA-BEI":"00b274", "UA-IPS":"00b274", "SEEPEX":"00b274", "SEEPEX":"00b274", "CROPEX":"00b274", "IBEX":"00b274", "SMP":"00b274", "PMUM":"00b274", "EXAA":"00b274", "SP":"00b274", "EEX":"00b274", "OTE":"00b274", "OKTE":"00b274", "OPCOM":"00b274", "SKUA":"00b274", "UASK":"00b274", "HUUA":"00b274", "UAHU":"00b274", "ROUA":"00b274", "UARO":"00b274",}

filename = path_to_scripts + 'Market_'+year+'.xlsx'
file_today = path_to_scripts + 'Market_'+year_today+'.xlsx'

if os.path.isfile(filename):
	wb = load_workbook(filename)
	writer = pd.ExcelWriter(filename, engine='openpyxl')
	writer.book = wb
else:
	wb = openpyxl.Workbook()
	wb.save(filename)
	wb = load_workbook(filename)
	writer = pd.ExcelWriter(filename, engine='openpyxl')
	writer.book = wb

#iterate through borders
for direction in direction_to_insert:
	if direction not in wb.sheetnames: 
		ws = wb.create_sheet(direction)
		try:
			ws.sheet_properties.tabColor = colour[direction]
		except:
			pass
		#add header for price
		ws.append(["Delivery Date", '00 - 01', '01 - 02', '02 - 03', '03 - 04', '04 - 05', '05 - 06', '06 - 07', '07 - 08', '08 - 09', '09 - 10', '10 - 11', '11 - 12', '12 - 13', '13 - 14', '14 - 15', '15 - 16', '16 - 17', '17 - 18', '18 - 19', '19 - 20', '20 - 21', '21 - 22', '22 - 23', '23 - 24', "Base", "Peak", "Avg"])

		#copy header for volume
		for col_num in range(1, 26):
			ws.cell(row=1, column=col_num+28).value = ws.cell(row=1, column=col_num).value
		
		#format header						
		for cell in ws[1:1]:
			if cell is not None:
				cell.font = Font(bold="True")
				cell.border = Border(bottom=Side(border_style="thick",color='FF000000'))
				cell.alignment=Alignment(vertical='center', horizontal='center')

		wb.save(filename)

	else:
		ws = wb[direction]

	#define number of next row
	row_tomorrow = ws.max_row + 1 
	row_today = ws.max_row

	for datum in datumi:
		ws = wb[direction]
		tomorrow = datetime.strptime(datum, '%Y-%m-%d')
		weekday = tomorrow.weekday()
		row_tomorrow = ws.max_row + 1 
		query_ems = "select direction, date, hour, auction_price, offered_capacity FROM aukcija where date = '%s'; " % datum
		df_ems = pd.read_sql(query_ems, con = db_connection)
		#get price and offered_capacity data
		df_direction = df_ems.loc[df_ems["direction"] == direction] 
		base = df_direction["auction_price"].mean()
		peak = [x for x in df_direction["auction_price"].values.tolist()]
		peak_dir = [peak[y] for y in range(7,20)]
		peak_pr = sum(peak_dir)/len(peak_dir)
		avg = df_direction["offered_capacity"].mean()
		df_dir_1 = df_direction.drop(columns=["direction","date"])
		df_dir_1 = df_dir_1.set_index('hour')	
		df_dir_2 = df_dir_1.T
		df_dir_2.insert(0,"Delivery Date", datum)
		df_dir_price = df_dir_2.iloc[0,:].values.tolist()
		df_dir_capacity = df_dir_2.iloc[1,:].values.tolist()


		#insert price		
		for col, val in enumerate(df_dir_price, start=1):
			ws.cell(row_tomorrow, column=col).value = val
			cell = ws.cell(row_tomorrow, column=col)
			if cell is not None:
				odd_row(cell)
				apply_color(cell, weekday)
		
		#insert capacity				
		for col, val in enumerate(df_dir_capacity, start=29):
			ws.cell(row_tomorrow, column=col).value = val
			cell = ws.cell(row_tomorrow, column=col)
			if cell is not None:
				odd_row(cell)
				apply_color(cell, weekday)	

		#insert base, peak, average
		base_cell = ws.cell(row=row_tomorrow, column=26)
		base_cell.value = base
		base_cell.fill = PatternFill("solid", fgColor="fff2cc")
		peak_cell = ws.cell(row=row_tomorrow, column=27)
		peak_cell.value = peak_pr
		peak_cell.fill = PatternFill("solid", fgColor="fbe5d6")
		avg_cell = ws.cell(row=row_tomorrow, column=28)
		avg_cell.value = avg
		avg_cell.fill = PatternFill("solid", fgColor="e2f0d9")
		for cell in [base_cell, peak_cell, avg_cell]:
			cell.number_format = '#,##0.00'
			cell.font = Font(name='Calibri', size=12)
			cell.alignment=Alignment(vertical='center', horizontal='center')

		#date_cells
		date_cells = [ws.cell(row=row_tomorrow, column=1), ws.cell(row=row_tomorrow, column=29)]
		for dc in date_cells:
			date_color(dc)
			
		
		wb.save(filename)

try:
	del wb['Sheet']
	wb.save(filename)
except:
	pass
	
db_connection.close()
# Close the Pandas Excel writer and output the Excel file.
#writer.save() 
