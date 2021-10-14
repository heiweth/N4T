import mysql.connector
from mysql.connector import Error
from mysql.connector import errorcode
import csv
from datetime import date, timedelta, datetime
import time
import pandas as pd
from sqlalchemy import create_engine
import pymysql
import openpyxl
from styles_sql import *
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from openpyxl import *
import os


#additional lists
areas_not = ["PMUM", "SMP", "HUPX", "OPCOM", "OTE", "OKTE"]
datumi = [d.date() for d in pd.date_range(start="2019-12-31",end="2021-04-19")]
# area, date, hour, price
#sqlEngine = create_engine('mysql+pymysql://root:network@localhost/N4T')
sqlEngine = create_engine('mysql+pymysql://cene:Babastamena22@192.168.100.251/cene')
db_connection = sqlEngine.connect()

for datum in datumi:
	today = datum
	tomorrow = today + timedelta(days=1)
	before = today + timedelta(days=-1)
	month = tomorrow.strftime("%m")
	year = tomorrow.strftime("%Y")
	day = tomorrow.strftime("%d")
	year_today = today.strftime("%Y")
	day_today = today.strftime("%d")
	week_today = today.weekday()
	weekday = tomorrow.weekday()
	week_before = before.weekday()
	print(tomorrow)
	#queries
	query = "SELECT area, price, volume FROM berza where date = '%s' order by area, hour;" % str(tomorrow)
	query_today = "SELECT area, price, volume FROM berza where date = '%s' order by area, hour;" % str(today)
	query_areas = "select distinct area from berza where date = '%s';" % str(tomorrow)
	query_areas_today = "select distinct area from berza where date = '%s';" % str(today)

	aukcija_query = "SELECT direction, hour, auction_price, offered_capacity FROM aukcija where date = '%s' order by direction, hour;" % str(tomorrow)
	aukcija_today_query = "SELECT direction, hour, auction_price, offered_capacity FROM aukcija where date = '%s'order by direction, hour;" % str(today)
	aukcija_before_query = "SELECT direction, hour, auction_price, offered_capacity FROM aukcija where date = '%s'order by direction, hour;" % str(before)

	aukcija_directions = "select distinct direction from aukcija where date = '%s';" % str(tomorrow)
	aukcija_today_directions = "select distinct direction from aukcija where date = '%s';" % str(today)
	aukcija_before_directions = "select distinct direction from aukcija where date = '%s';" % str(before)

	# read berza data and distinct areas for tomorrow
	df = pd.read_sql(query, con=db_connection)
	df_areas = pd.read_sql(query_areas, con=db_connection)
	areas = [row['area'] for index, row in df_areas.iterrows()]

	# read berza data and distinct areas for today
	df_today = pd.read_sql(query_today, con=db_connection)
	df_areas_today = pd.read_sql(query_areas_today, con=db_connection)
	areas_today = [row['area'] for index, row in df_areas_today.iterrows()]

	# read aukcija data and distinct directions for tomorrow
	df_aukcija = pd.read_sql(aukcija_query, con=db_connection)
	df_directions = pd.read_sql(aukcija_directions, con=db_connection)
	directions = [row['direction'] for index, row in df_directions.iterrows()]

	# read aukcija data and distinct directions for today
	df_today_aukcija = pd.read_sql(aukcija_today_query, con=db_connection)
	df_today_directions = pd.read_sql(aukcija_today_directions, con=db_connection)
	directions_today = [row['direction'] for index, row in df_today_directions.iterrows()]

	# read aukcija data and distinct directions for before
	df_before_aukcija = pd.read_sql(aukcija_before_query, con=db_connection)
	df_before_directions = pd.read_sql(aukcija_before_directions, con=db_connection)
	directions_before = [row['direction'] for index, row in df_before_directions.iterrows()]

#	db_connection.close()

	#define files
	path_to_templates = "/home/cene-admin/n4t_templates/"
	filename = path_to_templates + 'Market_2020.xlsx'
	file_today = path_to_templates + 'Market_' + year_today + '.xlsx'

	if os.path.isfile(filename):
		wb = load_workbook(filename)
		writer = pd.ExcelWriter(filename, engine='openpyxl')
		writer.book = wb

	else:
		wb = openpyxl.Workbook()
		wb.save(filename)
		wb = load_workbook(filename)
		writer = pd.ExcelWriter(filename, engine='openpyxl')
		writer.book = wb

	#for sht in wb.sheetnames:
	for sht in ["RSME", "MERS"]:
		writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
		ws = wb[sht]

		# define number of next row
		row_tomorrow = ws.max_row + 1
		row_today = ws.max_row
		row_yesterday = ws.max_row - 1

	#	row_tomorrow = ws.max_row
	#	row_today = ws.max_row - 1
	#	row_yesterday = ws.max_row - 2	

	# iterate through areas
		if sht in areas:

			# insert price data
			df_area = df.loc[df["area"] == sht]
			df_area1 = df_area.T
			df_area1.insert(0, "Delivery Date", tomorrow.strftime("%m/%d/%Y"))
			df_area1.set_index("Delivery Date")
			df_area1.iloc[[1]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startrow=row_today, startcol=0)

			base_price_cell = ws.cell(row=row_tomorrow, column=26)
			base_price_cell.value = "=AVERAGE(B" + str(row_tomorrow) + ":Y" + str(row_tomorrow) + ")"
			peak_price_cell = ws.cell(row=row_tomorrow, column=27)
			peak_price_cell.value = "=AVERAGE(J" + str(row_tomorrow) + ":U" + str(row_tomorrow) + ")"

			# insert volume data
			if sht not in areas_not:
				df_area1.iloc[[2]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startrow=row_today, startcol=29)

				base_volume_cell = ws.cell(row=row_tomorrow, column=55)
				base_volume_cell.value = "=AVERAGE(AE" + str(row_tomorrow) + ":BB" + str(row_tomorrow) + ")"
				peak_volume_cell = ws.cell(row=row_tomorrow, column=56)
				peak_volume_cell.value = "=AVERAGE(AM" + str(row_tomorrow) + ":AX" + str(row_tomorrow) + ")"

			# insert missing price data for yesterday
			if sht in ["PMUM","UA-BEI","UA-IPS"]:
				df_today_area = df_today.loc[df_today["area"] == sht]
				df_today_area1 = df_today_area.T
				df_today_area1.insert(0, "Delivery Date", today.strftime("%m/%d/%Y"))
				df_today_area1.iloc[[1]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startrow=row_yesterday, startcol=0)
				base_today_price_cell = ws.cell(row=row_today, column=26)
				base_today_price_cell.value = "=AVERAGE(B" + str(row_today) + ":Y" + str(row_today) + ")"
				peak_today_price_cell = ws.cell(row=row_today, column=27)
				peak_today_price_cell.value = "=AVERAGE(J" + str(row_today) + ":U" + str(row_today) + ")"

			if sht in ["UA-BEI","UA-IPS"]:
				df_today_area1.iloc[[2]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startrow=row_yesterday, startcol=29)

				base_today_volume_cell = ws.cell(row=row_today, column=55)
				base_today_volume_cell.value = "=AVERAGE(AE" + str(row_today) + ":BB" + str(row_today) + ")"
				peak_today_volume_cell = ws.cell(row=row_today, column=56)
				peak_today_volume_cell.value = "=AVERAGE(AM" + str(row_today) + ":AX" + str(row_today) + ")"


			# format date cells
			dc = ws.cell(row=row_tomorrow, column=1)
			date_style(dc)
			apply_color(dc, weekday)

			# format inserted rows
			for col in ws.iter_cols(min_row=row_tomorrow, min_col=2, max_row=row_tomorrow, max_col=25):
				for cell in col:
					row_style(cell)
					apply_color(cell, weekday)

				# format base/peak cells
				for bc in [base_price_cell, peak_price_cell]:
					row_style(bc)

			if sht not in areas_not:
				dc = ws.cell(row=row_tomorrow, column=30)
				date_style(dc)
				apply_color(dc, weekday)
				for col in ws.iter_cols(min_row=row_tomorrow, min_col=31, max_row=row_tomorrow, max_col=54):
					for cell in col:
						row_style(cell)
						apply_color(cell, weekday)

				# format base/peak cells
				for bc in [base_volume_cell, peak_volume_cell]:
					row_style(bc)

			if sht in ["PMUM", "UA-IPS", "UA-BEI"]:
				dc = ws.cell(row=row_today, column=1)
				date_style(dc)
				apply_color(dc, week_today)
				for col in ws.iter_cols(min_row=row_today, min_col=2, max_row=row_today, max_col=25):
					for cell in col:
						row_style(cell)
						apply_color(cell, week_today)

				# format base/peak cells
				for bc in [base_today_price_cell, peak_today_price_cell]:
					row_style(bc)

			if sht in ["UA-IPS", "UA-BEI"]:
				dc = ws.cell(row=row_today, column=30)
				date_style(dc)
				apply_color(dc, week_today)
				for col in ws.iter_cols(min_row=row_today, min_col=31, max_row=row_today, max_col=54):
					for cell in col:
						row_style(cell)
						apply_color(cell, week_today)

				# format base/peak cells
				for bc in [base_today_volume_cell, peak_today_volume_cell]:
					row_style(bc)
					
			wb.save(filename)

		
	# iterate through borders
		elif sht in directions:
			if ws.cell(row=row_today, column=2).value is None and sht in directions_today and sht in ["RSBA","BARS"]:
				# get price and offered_capacity data
				df_direction = df_today_aukcija.loc[df_today_aukcija["direction"] == sht]
				df_direction1 = df_direction.T
				df_direction1.insert(0, "Delivery Date", today.strftime("%m/%d/%Y"))
				df_direction1.iloc[[1]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False,
												 startrow=row_yesterday, startcol=0)
				df_direction1.iloc[[2]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False,
												 startrow=row_yesterday, startcol=28)

				base = df_direction["auction_price"].mean()
				peak = [x for x in df_direction["auction_price"].values.tolist()]
				peak_dir = [peak[y] for y in range(7, 20)]
				peak_pr = sum(peak_dir) / len(peak_dir)
				avg = df_direction["offered_capacity"].mean()

				# insert base, peak, average
				base_cell = ws.cell(row=row_today, column=26)
				base_cell.value = base
				peak_cell = ws.cell(row=row_today, column=27)
				peak_cell.value = peak_pr
				avg_cell = ws.cell(row=row_today, column=28)
				avg_cell.value = avg

				# format base, peak, average
				for cell in [base_cell, peak_cell, avg_cell]:
					no_border(cell)
				base_cell.fill = PatternFill("solid", fgColor="fff2cc")
				peak_cell.fill = PatternFill("solid", fgColor="fbe5d6")
				avg_cell.fill = PatternFill("solid", fgColor="e2f0d9")

				# format rows
				for col in ws.iter_cols(min_row=row_today, min_col=2, max_row=row_today, max_col=25):
					for cell in col:
						row_style(cell)
						apply_color(cell, week_today)

				for col in ws.iter_cols(min_row=row_today, min_col=30, max_row=row_today, max_col=53):
					for cell in col:
						row_style(cell)
						apply_color(cell, week_today)

				# date_cells
				date_cells = [ws.cell(row=row_today, column=1), ws.cell(row=row_today, column=29)]
				for dc in date_cells:
					date_style(dc)
					apply_color(cell, week_today)

			##################################################
			# get price and offered_capacity data for tomorrow
			df_direction = df_aukcija.loc[df_aukcija["direction"] == sht]
			df_direction1 = df_direction.T
			df_direction1.insert(0, "Delivery Date", tomorrow.strftime("%m/%d/%Y"))
			df_direction1.iloc[[2]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startrow=row_today, startcol=0)
			df_direction1.iloc[[3]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startrow=row_today, startcol=28)

			base = df_direction["auction_price"].mean()
			peak = [x for x in df_direction["auction_price"].values.tolist()]
			peak_dir = [peak[y] for y in range(7, 20)]
			peak_pr = sum(peak_dir) / len(peak_dir)
			avg = df_direction["offered_capacity"].mean()

			# insert base, peak, average
			base_cell = ws.cell(row=row_tomorrow, column=26)
			base_cell.value = base
			peak_cell = ws.cell(row=row_tomorrow, column=27)
			peak_cell.value = peak_pr
			avg_cell = ws.cell(row=row_tomorrow, column=28)
			avg_cell.value = avg

			# format base, peak, average
			for cell in [base_cell, peak_cell, avg_cell]:
				no_border(cell)
			base_cell.fill = PatternFill("solid", fgColor="fff2cc")
			peak_cell.fill = PatternFill("solid", fgColor="fbe5d6")
			avg_cell.fill = PatternFill("solid", fgColor="e2f0d9")

			# date_cells
			date_cells = [ws.cell(row=row_tomorrow, column=1), ws.cell(row=row_tomorrow, column=29)]
			for dc in date_cells:
				date_style(dc)
				apply_color(dc, weekday)

			# format rows
			for col in ws.iter_cols(min_row=row_tomorrow, min_col=2, max_row=row_tomorrow, max_col=25):
				for cell in col:
					row_style(cell)
					apply_color(cell, weekday)

			for col in ws.iter_cols(min_row=row_tomorrow, min_col=30, max_row=row_tomorrow, max_col=53):
				for cell in col:
					row_style(cell)
					apply_color(cell, weekday)

			wb.save(filename)


		else:
			print(sht)
			cell = ws.cell(row=row_tomorrow, column=1)
			cell.value = tomorrow.strftime("%m/%d/%Y")
			date_style(cell)
			apply_color(cell, weekday)

			wb.save(filename)

		
	try:
		del wb['Sheet']
		wb.save(filename)
	except:
		pass

# Close the Pandas Excel writer and output the Excel file.
# wb.save(filename)
