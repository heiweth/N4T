import mysql.connector
from mysql.connector import Error
from mysql.connector import errorcode
import csv
from datetime import date, timedelta, datetime
import time
import pandas as pd
from sqlalchemy import create_engine
import pymysql
import openpyxl
from styles_sanja import *
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from openpyxl import *
import os

path_to_scripts = "/home/sanja/N4T/"

today = date.today()
tomorrow = today + timedelta(days=1)
month = tomorrow.strftime("%m")
year = tomorrow.strftime("%Y")
day = tomorrow.strftime("%d")
month_slo = tomorrow.strftime("%B")
month_today = today.strftime("%B")
year_today = today.strftime("%Y")
day_today = today.strftime("%d")
week_today = today.weekday()
wee_day = tomorrow.strftime("%d")
weekday = tomorrow.weekday()

areas_not = ["Turkey", "Greece", "Hungary", "Romania", "Czech", "ELIX", "PHELIX"]
#, "Hungary", "Romania", "Czech", "ELIX", "PHELIX"
# area, date, hour, price
sqlEngine = create_engine('mysql+pymysql://root:network@localhost/N4T')
db_connection = sqlEngine.connect()

query = "SELECT area, date, hour, price, volume FROM berza where date = '%s';" % str(tomorrow)
query_areas = "select distinct area from berza where date = '%s';" % str(tomorrow)
query_not = "SELECT * FROM berza where date = '%s';" % str(today)

aukcija_query = "SELECT direction, date, hour, auction_price, offered_capacity FROM aukcija where date = '%s';" % str(tomorrow)
aukcija_directions = "select distinct direction from aukcija where date = '%s';" % str(tomorrow)

df = pd.read_sql(query, con = db_connection)
df_areas = pd.read_sql(query_areas, con = db_connection)
df_price_today = pd.read_sql(query_not, con = db_connection)
df_aukcija = pd.read_sql(aukcija_query, con = db_connection)
df_directions = pd.read_sql(aukcija_directions, con = db_connection)
directions = [row['direction'] for index,row in df_directions.iterrows()]
areas = [row['area'] for index,row in df_areas.iterrows()]
db_connection.close()

#tamnosiva
#ATHU HUAT HUHR HRHU HRRS RSHR SLOAT SIHR ATSLO HRSI GRIT ITGR CZDE(TenneT) DE(TenneT)CZ ATCZ CZAT

#zelena
#UA-BEI UA-IPS BRINDISI PUN SEEPEX CROPEX IBEX SMP PMUM EXAA SP EEX OTE OKTE OPCOM SKUA UASK HUUA UAHU ROUA UARO

#narandzasta
#HURS RSHU RORS RSRO BGRS RSBG MKRS RSMK

#base color
#ROBG BGRO

#peak color 
#BAHR HRBA MEBA BAME ALME MEAL ALGR GRAL GRTR TRGR MKGR GRMK

colour = {"BAHR":"fbe5d6" , "HRBA":"fbe5d6" , "MEBA":"fbe5d6" , "BAME":"fbe5d6" , "ALME":"fbe5d6" , "MEAL":"fbe5d6" , "ALGR":"fbe5d6" , "GRAL":"fbe5d6" , "GRTR":"fbe5d6" , "TRGR":"fbe5d6" , "MKGR":"fbe5d6" , "GRMK":"fbe5d6", "ATHU":"999999", "HUAT":"999999", "HUHR":"999999", "HRHU":"999999", "HRRS":"999999", "RSHR":"999999", "SLOAT":"999999", "SIHR":"999999", "ATSLO":"999999", "HRSI":"999999", "GRIT":"999999", "ITGR":"999999", "CZDE(TenneT)":"999999", "DE(TenneT)CZ":"999999", "ATCZ":"999999", "CZAT":"999999", "ROBG":"fff2cc", "BGRO":"fff2cc", "HURS":"fdb94d", "RSHU":"fdb94d", "RORS":"fdb94d", "RSRO":"fdb94d", "BGRS":"fdb94d", "RSBG":"fdb94d", "MKRS":"fdb94d", "RSMK":"fdb94d", "UA-BEI":"00b274", "UA-IPS":"00b274", "SEEPEX":"00b274", "SEEPEX":"00b274", "CROPEX":"00b274", "IBEX":"00b274", "SMP":"00b274", "PMUM":"00b274", "EXAA":"00b274", "SP":"00b274", "EEX":"00b274", "OTE":"00b274", "OKTE":"00b274", "OPCOM":"00b274", "SKUA":"00b274", "UASK":"00b274", "HUUA":"00b274", "UAHU":"00b274", "ROUA":"00b274", "UARO":"00b274",}

filename = path_to_scripts + 'Market_'+month_slo+'_'+year+'.xlsx'
file_today = path_to_scripts + 'Market_'+month_today+'_'+year+'.xlsx'

row_price = int(wee_day)+1
row_today = int(day_today)+1

if os.path.isfile(filename):
	wb = load_workbook(filename)

	print(filename)
	writer = pd.ExcelWriter(filename, engine='openpyxl')
	writer.book = wb
else:
	wb = openpyxl.Workbook()
	wb.save(filename)
	wb = load_workbook(filename)
	writer = pd.ExcelWriter(filename, engine='openpyxl')
	writer.book = wb

#iterate through areas
for area in areas:
	if area not in wb.sheetnames: 
		ws = wb.create_sheet(area)
		try:
			ws.sheet_properties.tabColor = colour[area]
		except:
			pass		
		#add header for price
		ws.append(["Delivery Date", '00 - 01', '01 - 02', '02 - 03', '03 - 04', '04 - 05', '05 - 06', '06 - 07', '07 - 08', '08 - 09', '09 - 10', '10 - 11', '11 - 12', '12 - 13', '13 - 14', '14 - 15', '15 - 16', '16 - 17', '17 - 18', '18 - 19', '19 - 20', '20 - 21', '21 - 22', '22 - 23', '23 - 24', "Base", "Peak"])
			
		#copy header for volume
		for col_num in range(1, 28):
			if area not in areas_not:	
				ws.cell(row=1, column=col_num+29).value = ws.cell(row=1, column=col_num).value

		#format header						
		for cell in ws[1:1]:
			if cell is not None:
				cell.font = Font(bold="True")
				cell.border = Border(bottom=Side(border_style="thick",color='FF000000'))
				cell.alignment=Alignment(vertical='center', horizontal='center')

		wb.save(filename)

	else:
		ws = wb[area]
		
	#get price and volume data
	if area not in areas_not:
		df_vol, df_price = read_mysql(df, area, tomorrow, areas_not)			
	else:
		df_price = read_mysql(df, area, tomorrow, areas_not)
		df_area = df_price_today.loc[df_price_today["area"] == area]
		df_pa_1 = df_area.drop(columns=["area","date"])
		df_pa_1 = df_pa_1.set_index('hour')	
		df_pa_2 = df_pa_1.T
		df_pa_2.insert(0,"Delivery Date", str(today))
		df_price_t = df_pa_2.iloc[0,:].values.tolist()
		
	#insert price data			
	for col, val in enumerate(df_price, start=1):
		ws.cell(row_price, column=col).value = val
		cell = ws.cell(row_price, column=col)
		if cell is not None:
			odd_row(cell)
			apply_color(cell, weekday, day)

	#insert volume data
	if area not in areas_not:
		for col, val in enumerate(df_vol, start=30):
			ws.cell(row_price, column=col).value = val		
			cell = ws.cell(row_price, column=col)
			if cell is not None:
				odd_row(cell)
				apply_color(cell, weekday, day)
					
	#insert missing price data for yesterday
	else:
		for col, val in enumerate(df_price_t, start=1):
			ws.cell(row = row_today, column=col).value = val
			cell = ws.cell(row_today, column=col)
			if cell is not None:
				odd_row(cell)
				apply_color(cell, weekday, day_today)			
			base_cell = ws.cell(row=row_today, column=26)
			base_cell.value = "=AVERAGE(B"+str(row_today)+":Y"+str(row_today)+")"
			peak_cell = ws.cell(row=row_today, column=27)
			peak_cell.value = "=AVERAGE(J"+str(row_today)+":U"+str(row_today)+")"

	#calculate base and peak values
	#format inserted cells
	for j in [(row_today,1,26,"B","Y",27,"J","U"),(row_price,1,26,"B","Y",27,"J","U"),(row_price,30,55,"AE","BB",56,"AM","AX")]:
		if (j[0] != row_today and area not in areas_not) or (area in areas_not and j[1] != 30) :
			cell_date = ws.cell(row=j[0], column=j[1])
			date_color(cell_date)
			base_cell = ws.cell(row=j[0], column=j[2])
			base_cell.value = "=AVERAGE("+j[3]+str(j[0])+":"+j[4]+str(j[0])+")"
			peak_cell = ws.cell(row=j[0], column=j[5])
			peak_cell.value = "=AVERAGE("+j[6]+str(j[0])+":"+j[7]+str(j[0])+")"

			for cell in [base_cell, peak_cell]:
				cell.number_format = '#,##0.00'
				cell.font = Font(name='Calibri', size=12)
				cell.alignment=Alignment(vertical='center', horizontal='center')
		
	wb.save(filename)


#iterate through borders
for direction in directions:
	if direction not in wb.sheetnames: 
		ws = wb.create_sheet(direction)
		try:
			ws.sheet_properties.tabColor = colour[direction]
		except:
			pass
		#add header for price
		ws.append(["Delivery Date", '00 - 01', '01 - 02', '02 - 03', '03 - 04', '04 - 05', '05 - 06', '06 - 07', '07 - 08', '08 - 09', '09 - 10', '10 - 11', '11 - 12', '12 - 13', '13 - 14', '14 - 15', '15 - 16', '16 - 17', '17 - 18', '18 - 19', '19 - 20', '20 - 21', '21 - 22', '22 - 23', '23 - 24', "Base", "Peak", "Avg"])

		#copy header for volume
		for col_num in range(1, 26):
			ws.cell(row=1, column=col_num+28).value = ws.cell(row=1, column=col_num).value
		
		#format header						
		for cell in ws[1:1]:
			if cell is not None:
				cell.font = Font(bold="True")
				cell.border = Border(bottom=Side(border_style="thick",color='FF000000'))
				cell.alignment=Alignment(vertical='center', horizontal='center')

		wb.save(filename)

	else:
		ws = wb[direction]
		
	#get price and offered_capacity data
	df_direction = df_aukcija.loc[df_aukcija["direction"] == direction] 
	base = df_direction["auction_price"].mean()
	peak = [x for x in df_direction["auction_price"].values.tolist()]
	peak_dir = [peak[y] for y in range(7,20)]
	peak_pr = sum(peak_dir)/len(peak_dir)
	avg = df_direction["offered_capacity"].mean()
	df_dir_1 = df_direction.drop(columns=["direction","date"])
	df_dir_1 = df_dir_1.set_index('hour')	
	df_dir_2 = df_dir_1.T
	df_dir_2.insert(0,"Delivery Date", str(tomorrow))
	df_dir_price = df_dir_2.iloc[0,:].values.tolist()
	df_dir_capacity = df_dir_2.iloc[1,:].values.tolist()
	
	#insert price		
	for col, val in enumerate(df_dir_price, start=1):
		ws.cell(row_price, column=col).value = val
		cell = ws.cell(row_price, column=col)
		if cell is not None:
			odd_row(cell)
			apply_color(cell, weekday, day)
	
	#insert capacity				
	for col, val in enumerate(df_dir_capacity, start=29):
		ws.cell(row_price, column=col).value = val
		cell = ws.cell(row_price, column=col)
		if cell is not None:
			odd_row(cell)
			apply_color(cell, weekday, day)	

	#insert base, peak, average
	base_cell = ws.cell(row=row_price, column=26)
	base_cell.value = base
	base_cell.fill = PatternFill("solid", fgColor="fff2cc")
	peak_cell = ws.cell(row=row_price, column=27)
	peak_cell.value = peak_pr
	peak_cell.fill = PatternFill("solid", fgColor="fbe5d6")
	avg_cell = ws.cell(row=row_price, column=28)
	avg_cell.value = avg
	avg_cell.fill = PatternFill("solid", fgColor="e2f0d9")
	for cell in [base_cell, peak_cell, avg_cell]:
		cell.number_format = '#,##0.00'
		cell.font = Font(name='Calibri', size=12)
		cell.alignment=Alignment(vertical='center', horizontal='center')

	date_cells = [ws.cell(row=row_price, column=1), ws.cell(row=row_price, column=29)]
	for dc in date_cells:
		date_color(dc)
		
	



	wb.save(filename)

try:
	std=wb.get_sheet_by_name('Sheet')
	wb.remove_sheet(std)
	wb.save(filename)
except:
	pass
	

# Close the Pandas Excel writer and output the Excel file.
#writer.save() 
