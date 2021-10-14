import mysql.connector
from sqlalchemy import create_engine
import csv
from datetime import date, timedelta, datetime
import time
import pandas as pd
import numpy as np
import pymysql
import openpyxl
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from openpyxl import *
import os
import xlrd
import xlwt
from xlutils.copy import copy

path_to_templates = "/home/cene-admin/n4t_templates/"

today = date.today()
tomorrow = today + timedelta(days=1)

auction_results = {"EMS":("RSMK","MKRS","RSRO","RORS","RSHU","HURS"),"NOS":("BARS","RSBA"),
				"JAO":("ATHU","HUAT","HRHU","HUHR","BGGR","GRBG","HRRS","RSHR", "BGRS", "RSBG"),
				"SEECAO":("BAHR","HRBA","MEBA","BAME","ALME","MEAL","ALGR","GRAL","GRTR","TRGR","MKGR","GRMK"),
				"TRE":("ROBG","BGRO")}

exchange = ["BRINDISI", "EEX", "EXAA", "HUPX", "IBEX", "OKTE", "SMP", "SEEPEX", "OTE", "OPCOM", "SP", "CROPEX"]


# area, date, hour, price
#sqlEngine = create_engine('mysql+pymysql://root:network@localhost/N4T')
sqlEngine = create_engine('mysql+pymysql://cene:Babastamena22@192.168.100.145/cene')
db_connection = sqlEngine.connect()


for key in auction_results.keys():
#for key in ["EMS"]:
	filename  = path_to_templates + "auction_results_template_V2-" + key + ".xlsx"
	wb = load_workbook(filename)
	writer = pd.ExcelWriter(filename, engine='openpyxl')
	writer.book = wb
	writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
	ws = wb['Sheet1']
	values = auction_results[key]
	df_prep = pd.DataFrame(columns=["direction", "date","hour","source","sink","auction_price"])
	for value in values:
		query = "select direction, date, hour, source, sink, auction_price from aukcija where date = '%s' and direction = '%s';" % (str(tomorrow), value)
		df_query = pd.read_sql(query, con=db_connection)
		df_prep = df_prep.append(df_query,ignore_index = True)

	unibord = np.unique(np.array(df_prep["direction"]))
	diff = set(set(values)-set(unibord))
	df_prep = df_prep.drop(columns="direction")
	if key == "TRE":
		ws["A3"] = ws["A4"] = ws["B3"] = ws["B4"] = tomorrow.strftime("%d.%m.%Y.")
		wb.save(filename)
	try:
		d11 = df_prep.pivot_table(index=["date","source","sink"],columns=["hour"], values="auction_price")
	except:
		continue
	d12 = d11.reset_index()
	d12.insert(0,"Date From", df_prep["date"])

	print(d12)
	d12.columns=["Date from", "Date to", "Source", "Sink", "H01", "H02","H03", "H04", "H05", "H06", "H07", "H08", "H09", "H10", "H11",
						"H12", "H13", "H14", "H15", "H16", "H17", "H18", "H19", "H20", "H21", "H22", "H23","H24"]

	d12.to_excel(writer, sheet_name="Sheet1", float_format="%.2f",startrow=2, index=False, header=False)

	wb.save(filename)

	max_row = ws.max_row

#	for value in diff:
#		ws["A3"] = ws["A4"] = ws["B3"] = ws["B4"] = tomorrow.strftime("%d.%m.%Y.")

	for row in ws.iter_rows(min_row=3, min_col=1, max_col=2):
		for cell in row:
			try:
				cell.value = cell.value.strftime("%d.%m.%Y.")
			except:
				cell.value = tomorrow.strftime("%d.%m.%Y.")
			finally:
				continue
				
	wb.save(filename)

for exc in exchange:
	#create and load wb
	filename = path_to_templates + "exchange_template_" + exc + ".xls"
	rb = xlrd.open_workbook(filename)
	wb = copy(rb)
	ws = wb.get_sheet("Prices")
	#get the data
	query = "select date, hour, price from berza where date = '%s' and area = '%s';" % (str(tomorrow), exc)
	df_query = pd.read_sql(query, con=db_connection)

	try:
		d11 = df_query.pivot_table(index=["date"],columns=["hour"], values="price")
	except:
		continue
	
	d12 = d11.reset_index()
	d13 = d12.values
	d13[0][0] = d13[0][0].strftime("%d.%m.%Y.")

	for i in range(len(d13[0])):
		ws.write(2, i, d13[0][i])

	wb.save(filename)

db_connection.close()

