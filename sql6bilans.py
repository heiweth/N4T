import mysql.connector
from mysql.connector import Error
from mysql.connector import errorcode
import csv
from datetime import date, timedelta, datetime
import time
import pandas as pd
from sqlalchemy import create_engine
import pymysql
import openpyxl
from styles_sql import *
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from openpyxl import *
import os

today = date.today()
tomorrow = today + timedelta(days=1)

#additional lists
berza_list = ["HUPX", "OPCOM", "SMP", "OKTE", "CROPEX", "IBEX", "SEEPEX", "UA-BEI"]
aukcija_list = ["HURS", "HUHR", "BARS", "RSBA", "RSHU", "ALGR", "MKGR", "ROBG", "BGRO", "GRAL",  "GRMK", "ALME", "MEAL", "BAME", "RSHR", "MEBA", "RSRO", "RORS", "BGRS","MERS", "HRBA", "RSMK", "MKRS",
 "BAHR", "RSBG", "HRHU", "HRRS", "BGGR", "GRBG", "ITME"]


#get data from db
#sqlEngine = create_engine('mysql+pymysql://root:network@localhost/N4T')
sqlEngine = create_engine('mysql+pymysql://cene:Babastamena22@192.168.100.145/cene')
db_connection = sqlEngine.connect()

#queries
berza_query = "SELECT area, hour, price FROM berza where date = '%s' order by hour;" % str(tomorrow)
aukcija_query = "SELECT direction, hour, auction_price FROM aukcija where date = '%s' order by direction, hour;" % str(tomorrow)

# read berza data and distinct areas for tomorrow
df_berza = pd.read_sql(berza_query, con=db_connection)

# read aukcija data and distinct directions for tomorrow
df_aukcija = pd.read_sql(aukcija_query, con=db_connection)

db_connection.close()

#define files
path_to_docs = "/home/cene-admin/n4t_docs/"
path_to_templates = "/home/cene-admin/n4t_templates/"
file_bilans = path_to_templates + 'Bilans_' + str(tomorrow) + '.xlsx'

wbilans = openpyxl.Workbook()
wbilans.save(file_bilans)
wbilans = load_workbook(file_bilans)
writer = pd.ExcelWriter(file_bilans, engine='openpyxl')
writer.book = wbilans

ws = wbilans.create_sheet("DATA")
writer.sheets = dict((ws.title, ws) for ws in wbilans.worksheets)

for area in berza_list:
	maxrow = ws.max_row
	df11 = df_berza.loc[df_berza["area"]==area]
	if df11.empty:
		for x in range(1,25):
			df11 = df11.append({"area" : area, "hour" : x }, ignore_index = True)
			df11 = df11.fillna(" ")
	df12 = df11.T
	df12.iloc[[2]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startcol=1, startrow = maxrow)
	ws.cell(row=maxrow+1, column=1).value	 = area	

	wbilans.save(file_bilans)


for border in aukcija_list:
	maxrow = ws.max_row
	df21 = df_aukcija.loc[df_aukcija["direction"]==border]
	if df21.empty : 
		for x in range(1,25):
			df21 = df21.append({"direction" : border, "hour" : x }, ignore_index = True)
			df21 = df21.fillna(" ")
	
	df22 = df21.T 
	df22.iloc[[2]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startcol=1, startrow = maxrow)
	ws.cell(row=maxrow+1, column=1).value	 = border
	
	wbilans.save(file_bilans)

ws.delete_rows(1)
wbilans.save(file_bilans)
time.sleep(2)

for row in ws.iter_rows():
	if row[0].value in berza_list or row[0].value == "CROPEX":
		for cell in row:
			berza_row(cell)
	else:
		for cell in row:
			aukcija_row(cell)
	base_price_cell = ws.cell(row=row[0].row, column=26)
	base_price_cell.value = "=AVERAGE(B" + str(row[0].row) + ":Y" + str(row[0].row) + ")"
	no_border(base_price_cell)
	row[0].font = Font(bold="True")
	row[0].border = Border(right=Side(border_style="thin", color='FF000000'), bottom=Side(border_style="thin", color='FF000000'),left=Side(border_style="thin", color='FF000000'))	
	
try:
	del wbilans['Sheet']
	wbilans.save(file_bilans)
except:
	pass
	
wbilans.save(file_bilans)

