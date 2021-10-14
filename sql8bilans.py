import mysql.connector
from mysql.connector import Error
from mysql.connector import errorcode
import csv
from datetime import date, timedelta, datetime
import time
import pandas as pd
from sqlalchemy import create_engine
import pymysql
import openpyxl
from styles_sql import *
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from openpyxl import *
import os

today = date.today()
tomorrow = today + timedelta(days=1)
day = today 

#additional lists
berza_list = ["HUPX", "OPCOM", "SMP", "OKTE", "BRINDISI", "IBEX", "SEEPEX", "UA-BEI"]
aukcija_list = ["HURS", "HUHR", "BARS", "RSBA", "RSHU", "ALGR", "MKGR", "ROBG", "BGRO", "GRAL",  "GRMK", "ALME", "MEAL", "BAME", "RSHR", "MEBA", "RSRO", "RORS", "BGRS","MERS", "HRBA", "RSMK", "MKRS",
 "BAHR", "RSBG", "HRHU", "HRRS", "BGGR", "GRBG", "ITME"]

#get data from db
#sqlEngine = create_engine('mysql+pymysql://root:network@localhost/N4T')
sqlEngine = create_engine('mysql+pymysql://cene:Babastamena22@192.168.100.145/cene')
db_connection = sqlEngine.connect()

#queries
berza_query = "SELECT area, hour, price FROM berza where date = '%s' order by hour;" % str(day)
aukcija_query = "SELECT direction, hour, auction_price FROM aukcija where date = '%s' order by direction, hour;" % str(day)

# read berza data and distinct areas for day
df_berza = pd.read_sql(berza_query, con=db_connection)

# read aukcija data and distinct directions for day
df_aukcija = pd.read_sql(aukcija_query, con=db_connection)

db_connection.close()

#define files
path_to_docs = "/home/cene-admin/n4t_docs/"
path_to_templates = "/home/cene-admin/n4t_templates/"
file_bilans = path_to_templates + 'Bilans_' + str(day) + '.xlsx'

try:
	wbilans = load_workbook(file_bilans)
	writer = pd.ExcelWriter(file_bilans, engine='openpyxl')
	writer.book = wbilans
except:
	sys.exit()

#define dfs
df_area = pd.DataFrame(columns=["area","hour","price"])
df_border = pd.DataFrame(columns=["direction","hour","auction_price"])

ws = wbilans["DATA"]
writer.sheets = dict((ws.title, ws) for ws in wbilans.worksheets)

dict_row = {"SMP":2, "UA-BEI":7}

for area in dict_row.keys():
	df11 = df_berza.loc[df_berza["area"]==area]
	if df11.empty:
		for x in range(1,25):
			df11 = df11.append({"area" : area, "hour" : x }, ignore_index = True)
			df11 = df11.fillna(" ")
	df12 = df11.T
	df12.iloc[[2]].to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, index=False, startcol=1, startrow = dict_row[area])
	ws.cell(row=dict_row[area]+1, column=1).value = area	

try:
	
	del wbilans['Sheet']
	wbilans.save(file_bilans)
except:
	pass
	
wbilans.save(file_bilans)
wbilans.close()

