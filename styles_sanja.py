from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from datetime import datetime

def apply_color(cell, week):
	if week > 4 :
		cell.fill = PatternFill("solid", fgColor="cccccc")			
	elif int(week) % 2 == 1 :
		cell.fill = PatternFill("solid", fgColor="b7d3ee")
	return(cell)
	
def date_color(cell):
	cell.value = cell.value.strftime("%m/%d/%Y")
	return(cell)

def odd_row(cell):
	cell.font = Font(name='Calibri', size=12)
	cell.number_format = '#,##0.00'
	cell.alignment=Alignment(vertical='center', horizontal='center')
	cell.border = Border(right=Side(border_style="thin",color='4472c4'), bottom=Side(border_style="thin",color='4472c4'))
	return(cell)
	
def add_price(ws, df_price, day, weekday):
	for col, val in enumerate(df_price, start=1):
		ws.cell(row=int(day), column=col).value = val
		cell = ws.cell(row=int(day), column=col)
		odd_row(cell)
		apply_color(cell, weekday, day)
		return(ws.cell.value)

def add_volume(ws, df_vol, day, weekday):
	for col, val in enumerate(df_vol, start=1):
		ws.cell(row=int(day)+35, column=col).value = val		
		cell = ws.cell(row=int(day)+35, column=col)
		
		odd_row(cell)
		apply_color(cell, weekday, day)
		return(ws.cell.value)

def format_date(ws, day):
	for i in [int(day), int(day)+35]:
		if ws.cell(i, column=2).value is not None:
			cell_date = ws.cell(i, column=1)
			date_color(cell_date)
			return(cell_date)
	
def read_mysql(df, area, tomorrow, areas_not):
	df_area = df.loc[df["area"] == area] 
	base_price = df_area["price"].mean()
	peak_prices = [x for x in df_area["price"].values.tolist()]
	peak_p = [peak_prices[y] for y in range(7,20)]
	peak_price = sum(peak_p)/len(peak_p)
	if area not in areas_not:
		base_volume = df_area["volume"].mean()
		peak_volumes = [x for x in df_area["volume"].values.tolist()]
		peak_v = [peak_volumes[y] for y in range(7,20)]
		peak_volume = sum(peak_v)/len(peak_v)
	
	df_area_1 = df_area.drop(columns=["area","date"])
	df_area_1 = df_area_1.set_index('hour')	
	df_area_2 = df_area_1.T
	df_area_2.insert(0,"Delivery Date", str(tomorrow))
	df_price = df_area_2.iloc[0,:].values.tolist()
	if area not in areas_not:
		df_vol = df_area_2.iloc[1,:].values.tolist()
	
	if area not in areas_not:
		return(df_vol, df_price)
	else: 
		return(df_price)
