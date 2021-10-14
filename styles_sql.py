from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from datetime import datetime

def berza_row(cell):
	cell.font=Font(name='Calibri', size=12)
	cell.alignment=Alignment(vertical='center', horizontal='center')
	cell.fill = PatternFill("solid", fgColor="FFE5CA")
	
def aukcija_row(cell):
	cell.font=Font(name='Calibri', size=12)
	cell.alignment=Alignment(vertical='center', horizontal='center')
	cell.fill = PatternFill("solid", fgColor="DEEBF7")
	
def apply_color(cell, week):
	if week > 4 :
		cell.fill = PatternFill("solid", fgColor="cccccc")
	elif int(week) % 2 == 1 :
		cell.fill = PatternFill("solid", fgColor="b7d3ee")
	return(cell)

# styles
def date_style(cell):
	cell.font=Font(name='Calibri', size=12, bold="True")
	cell.border=Border(right=Side(border_style="thin", color='FF000000'), bottom=Side(border_style="thin", color='FF000000'),
					   left=Side(border_style="thin", color='FF000000'))
	return(cell)

def row_style(cell):
	cell.number_format='#,##0.00'
	cell.alignment=Alignment(vertical='center', horizontal='center')
	cell.font=Font(name='Calibri', size=12)
	cell.border=Border(right=Side(border_style="thin", color='4472c4'), bottom=Side(border_style="thin", color='4472c4'))
	return(cell)

def no_border(cell):
	cell.number_format='#,##0.00'
	cell.font=Font(name='Calibri', size=12)
	cell.alignment=Alignment(vertical='center', horizontal='center')
	return(cell)
