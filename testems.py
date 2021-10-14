import mysql.connector
from mysql.connector import Error
from mysql.connector import errorcode
import csv
from datetime import date, timedelta, datetime
import time
import pandas as pd
from sqlalchemy import create_engine
import pymysql
import openpyxl
from styles_sql import *
from styles_sanja import *
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from openpyxl import *
import os

path_to_scripts = "/home/cene-admin/n4t/"
path_to_templates = "/home/cene-admin/n4t_templates/"

datumi = [d.date() for d in pd.date_range(start="2021-01-03",end="2021-04-19")]
#datumi = ["2020-07-27","2020-07-28"]
areas_not = ["PMUM", "SMP", "HUPX", "OPCOM", "OTE", "ELIX", "PHELIX"]
directions_to_insert = ["MERS", "RSME"]
area_to_insert = []


today = date.today()

# area, date, hour, price
#sqlEngine = create_engine('mysql+pymysql://root:network@localhost/N4T')
sqlEngine = create_engine('mysql+pymysql://cene:Babastamena22@192.168.100.251/cene')
db_connection = sqlEngine.connect()


filename = path_to_templates + 'Market_2020.xlsx'
#file_today = path_to_templates + 'Market_'+year_today+'.xlsx'

if os.path.isfile(filename):
	wb = load_workbook(filename)
	writer = pd.ExcelWriter(filename, engine='openpyxl', date_format="mm/dd/YYYY")
	writer.book = wb
else:
	wb = openpyxl.Workbook()
	wb.save(filename)
	wb = load_workbook(filename)
	writer = pd.ExcelWriter(filename, engine='openpyxl', date_format="mm/dd/YYYY")
	writer.book = wb

#iterate through borders

	#define number of next row
	 
#	row_today = ws.max_row

for datum in datumi:
	print(datum)
#		new_datum = datetime.strptime(datum, "%Y-%m-%d")
	weekday = datum.weekday()

	query_ems = "SELECT direction, date, hour, auction_price, offered_capacity FROM aukcija where date = '%s';" % datum
	df_ems = pd.read_sql(query_ems, con = db_connection)

	for direction in directions_to_insert:
		ws = wb[direction]
		writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
		row_tomorrow = ws.max_row + 1
		row_today = ws.max_row
		row_yesterday = ws.max_row - 1

		#get price and offered_capacity data
		df_direction = df_ems.loc[df_ems["direction"] == direction]
		df_direction.drop(columns="direction")
		df_pivot_price = df_direction.pivot_table(index=["date"],columns=["hour"], values=["auction_price", "offered_capacity"])
		df_pivot_capacity = df_direction.pivot_table(index=["date"],columns=["hour"], values=["offered_capacity"])
		df_pivot_price.to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, startrow=row_yesterday, startcol=0)
		df_pivot_capacity.to_excel(writer, sheet_name=ws.title, float_format="%.2f", header=False, startrow=row_yesterday, startcol=28)


		base = df_direction["auction_price"].mean()
		peak = [x for x in df_direction["auction_price"].values.tolist()]
		peak_dir = [peak[y] for y in range(7, 20)]
		peak_pr = sum(peak_dir) / len(peak_dir)
		avg = df_direction["offered_capacity"].mean()

		# insert base, peak, average
		base_cell = ws.cell(row=row_tomorrow, column=26)
		base_cell.value = base
		peak_cell = ws.cell(row=row_tomorrow, column=27)
		peak_cell.value = peak_pr
		avg_cell = ws.cell(row=row_tomorrow, column=28)
		avg_cell.value = avg

		# format base, peak, average
		for cell in [base_cell, peak_cell, avg_cell]:
			no_border(cell)
		base_cell.fill = PatternFill("solid", fgColor="fff2cc")
		peak_cell.fill = PatternFill("solid", fgColor="fbe5d6")
		avg_cell.fill = PatternFill("solid", fgColor="e2f0d9")

		# date_cells
		date_cells = [ws.cell(row=row_tomorrow, column=1), ws.cell(row=row_tomorrow, column=29)]
		for dc in date_cells:
			date_style(dc)
			apply_color(dc, weekday)
			date_color(dc)

		# format rows
		for col in ws.iter_cols(min_row=row_tomorrow, min_col=2, max_row=row_tomorrow, max_col=25):
			for cell in col:
				row_style(cell)
				apply_color(cell, weekday)

		for col in ws.iter_cols(min_row=row_tomorrow, min_col=30, max_row=row_tomorrow, max_col=53):
			for cell in col:
				row_style(cell)
				apply_color(cell, weekday)

				
		
		wb.save(filename)

try:
	del wb['Sheet']
	wb.save(filename)
except:
	pass
	
db_connection.close()
# Close the Pandas Excel writer and output the Excel file.
#writer.save() 
